import openpyxl
import json
from run_scrapers import scrape_all

def load_config():
    with open("config.json", "r") as config_file:
        config = json.load(config_file)
    return config

if __name__ == "__main__":
    # Load configuration
    config = load_config()
    excel_file_path = config["excel_file_path"]

    # Get all scraped data
    metrics = scrape_all()

    # Open excel file
    #excel_file_path = r"C:\Users\MarcoHui\Knightsbridge Asset Management, LLC\Knightsbridge - Documents\Portfolio Management\Kurt's Model Data\Kurt's model data.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook["Kurt's models"]
    
    # Find the last filled row
    last_row = sheet.max_row + 1

    # Write in data
    for i, j in zip(range(1, 8), metrics):
        sheet.cell(row=last_row, column=i, value=j)
    
    # Save changes
    workbook.save(excel_file_path)

    print(f'Data for {metrics[0]} added!')